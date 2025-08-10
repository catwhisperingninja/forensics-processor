import os
import glob
import openpyxl
from datetime import datetime

# Find the latest Excel file in the visualizations directory
files = glob.glob("test_default_workflow/visualizations/comprehensive_report_*.xlsx")
if not files:
    print("No Excel files found!")
    exit(1)

# Get the latest file by creation time
latest_file = max(files, key=os.path.getctime)
print(f"Opening latest file: {latest_file}")

# Open the workbook and list sheets
try:
    wb = openpyxl.load_workbook(latest_file)
    print(f"Successfully opened the workbook!")
    print(f"Sheets in workbook: {wb.sheetnames}")

    # Print some basic statistics about each sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"  - Dimensions: {sheet.dimensions}")
        print(f"  - Max row: {sheet.max_row}")
        print(f"  - Max column: {sheet.max_column}")

        # Print first few cells of data as a sample
        if sheet.max_row > 1 and sheet.max_column > 1:
            print("  - Sample data:")
            for r in range(1, min(5, sheet.max_row + 1)):
                row_data = []
                for c in range(1, min(5, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=r, column=c).value
                    row_data.append(str(cell_value)[:20] if cell_value else "None")
                print(f"    Row {r}: {row_data}")

except Exception as e:
    print(f"Error opening workbook: {str(e)}")