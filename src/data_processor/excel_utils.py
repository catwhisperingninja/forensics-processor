"""
Excel Utilities for advanced Excel file manipulation.
"""
import os
from typing import List, Dict, Any, Tuple, Optional, Union
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    """Class for Excel formatting operations."""

    # Color presets
    COLORS = {
        'light_blue': 'B4C6E7',
        'light_green': 'C6E0B4',
        'light_red': 'F8CECC',
        'light_yellow': 'FFF2CC',
        'light_orange': 'FFE699',
        'medium_blue': '8EA9DB',
        'medium_green': '97D077',
        'medium_red': 'F19C99',
        'header_blue': '4472C4',
        'header_green': '70AD47',
    }

    def __init__(self, file_path: str):
        """Initialize formatter with an Excel file path."""
        self.file_path = file_path
        # Always create a new workbook
        self.workbook = openpyxl.Workbook()
        # Save the workbook
        self.workbook.save(file_path)

    def create_sheet(self, sheet_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
        """Create a new sheet if it doesn't exist."""
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        return self.workbook.create_sheet(sheet_name)

    def df_to_excel_table(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        table_name: str,
        start_cell: str = 'A1',
        style: str = 'TableStyleMedium9'
    ) -> None:
        """
        Convert DataFrame to formatted Excel table.

        Args:
            df: Pandas DataFrame to convert
            sheet_name: Target worksheet name
            table_name: Name for the Excel table
            start_cell: Starting cell for table (e.g., 'A1')
            style: Excel table style to apply
        """
        sheet = self.create_sheet(sheet_name)

        # Starting row and column
        row, col = self._cell_to_coordinates(start_cell)

        # Write headers
        for c, header in enumerate(df.columns):
            sheet.cell(row=row, column=col+c).value = header

        # Write data
        for r, data_row in enumerate(df.values):
            for c, value in enumerate(data_row):
                sheet.cell(row=row+r+1, column=col+c).value = value

        # Calculate table dimensions
        end_row = row + len(df)
        end_col = col + len(df.columns) - 1
        table_dims = f"{start_cell}:{get_column_letter(end_col)}{end_row}"

        # Create Excel table
        table = Table(displayName=table_name, ref=table_dims)
        table.tableStyleInfo = TableStyleInfo(
            name=style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        # Remove any existing table with the same name
        tables_to_remove = []
        for existing_table_name in list(sheet.tables.keys()):
            if existing_table_name == table_name:
                tables_to_remove.append(existing_table_name)

        for table_name_to_remove in tables_to_remove:
            sheet.tables.pop(table_name_to_remove)

        sheet.add_table(table)
        self.workbook.save(self.file_path)

    def apply_header_style(
        self,
        sheet_name: str,
        header_range: str,
        bold: bool = True,
        bg_color: str = 'header_blue',
        font_color: str = 'FFFFFF',
        font_size: int = 11
    ) -> None:
        """Apply styling to header cells."""
        sheet = self.workbook[sheet_name]
        bg_color = self.COLORS.get(bg_color, bg_color)

        # Define style
        font = Font(bold=bold, color=font_color, size=font_size)
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        alignment = Alignment(horizontal='center', vertical='center')

        # Apply style
        for row in sheet[header_range]:
            for cell in row:
                cell.font = font
                cell.fill = fill
                cell.alignment = alignment

        self.workbook.save(self.file_path)

    def apply_conditional_formatting(
        self,
        sheet_name: str,
        cell_range: str,
        rule_type: str,
        formula: Optional[str] = None,
        min_val: Optional[float] = None,
        max_val: Optional[float] = None,
        colors: Optional[List[str]] = None
    ) -> None:
        """
        Apply conditional formatting to a range of cells.

        Args:
            sheet_name: Target worksheet
            cell_range: Cell range to format (e.g., 'A2:D20')
            rule_type: 'color_scale', 'cell_is', or 'formula'
            formula: Formula for 'formula' type
            min_val, max_val: Values for 'cell_is' type
            colors: List of color hex codes
        """
        sheet = self.workbook[sheet_name]

        if rule_type == 'color_scale':
            colors = colors or ['F8696B', 'FFEB84', '63BE7B']  # Red to Green
            rule = ColorScaleRule(start_type='min', start_color=colors[0],
                                 mid_type='percentile', mid_value=50, mid_color=colors[1],
                                 end_type='max', end_color=colors[2])

        elif rule_type == 'cell_is':
            colors = colors or ['F8696B']  # Default red
            if min_val is None or max_val is None:
                raise ValueError("min_val and max_val required for 'cell_is' rule")

            rule = CellIsRule(
                operator='between',
                formula=[str(min_val), str(max_val)],
                stopIfTrue=True,
                fill=PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
            )

        elif rule_type == 'formula':
            colors = colors or ['F8696B']  # Default red
            if formula is None:
                raise ValueError("formula required for 'formula' rule")

            rule = FormulaRule(
                formula=[formula],
                stopIfTrue=True,
                fill=PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
            )
        else:
            raise ValueError(f"Unsupported rule_type: {rule_type}")

        sheet.conditional_formatting.add(cell_range, rule)
        self.workbook.save(self.file_path)

    def autofit_columns(self, sheet_name: str) -> None:
        """Autosize columns based on content."""
        sheet = self.workbook[sheet_name]

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width

        self.workbook.save(self.file_path)

    def add_border(self, sheet_name: str, cell_range: str, border_style: str = 'thin') -> None:
        """Add borders to a range of cells."""
        sheet = self.workbook[sheet_name]
        border = Border(
            left=Side(style=border_style),
            right=Side(style=border_style),
            top=Side(style=border_style),
            bottom=Side(style=border_style)
        )

        for row in sheet[cell_range]:
            for cell in row:
                cell.border = border

        self.workbook.save(self.file_path)

    def _cell_to_coordinates(self, cell: str) -> Tuple[int, int]:
        """Convert cell reference (e.g., 'A1') to row, col coordinates."""
        # Separate letter and number parts
        col_str = ''.join(filter(str.isalpha, cell))
        row_str = ''.join(filter(str.isdigit, cell))

        # Convert column letters to number (A=1, B=2, etc.)
        col_num = 0
        for char in col_str:
            col_num = col_num * 26 + (ord(char.upper()) - ord('A') + 1)

        return int(row_str), col_num

    def highlight_cells(
        self,
        sheet_name: str,
        cell_range: str,
        bg_color: str,
        font_color: str = '000000'
    ) -> None:
        """Highlight specific cells with background color."""
        sheet = self.workbook[sheet_name]
        bg_color = self.COLORS.get(bg_color, bg_color)

        font = Font(color=font_color)
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        for row in sheet[cell_range]:
            for cell in row:
                cell.font = font
                cell.fill = fill

        self.workbook.save(self.file_path)

    def save(self) -> None:
        """Save the workbook."""
        self.workbook.save(self.file_path)

    def close(self) -> None:
        """Close the workbook."""
        self.workbook.close()


def create_summary_report(data_file: str, output_file: str, sheet_name: str = 'Summary') -> None:
    """
    Create a summary Excel report from a CSV or Excel file.

    Args:
        data_file: Path to source data file (CSV or Excel)
        output_file: Path to save the formatted Excel report
        sheet_name: Name of the summary sheet
    """
    # Load data
    if data_file.endswith('.csv'):
        df = pd.read_csv(data_file)
    else:
        df = pd.read_excel(data_file)

    # Create formatter
    formatter = ExcelFormatter(output_file)

    # Create summary sheet with table
    formatter.df_to_excel_table(
        df=df,
        sheet_name=sheet_name,
        table_name='SummaryTable',
        start_cell='A1'
    )

    # Apply styling
    formatter.apply_header_style(
        sheet_name=sheet_name,
        header_range='A1:' + get_column_letter(len(df.columns)) + '1'
    )

    # Auto-size columns
    formatter.autofit_columns(sheet_name)

    # Add borders
    formatter.add_border(
        sheet_name=sheet_name,
        cell_range='A1:' + get_column_letter(len(df.columns)) + str(len(df) + 1)
    )

    formatter.save()
    formatter.close()

    return output_file