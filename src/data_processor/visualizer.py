"""
Data Visualization module for creating charts and visualizations in Excel.
"""
import os
from typing import List, Dict, Any, Tuple, Optional, Union
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.chart import (
    BarChart, LineChart, PieChart, ScatterChart, BubbleChart,
    Reference, Series
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.utils import get_column_letter


class ExcelChartCreator:
    """Class for creating and embedding charts in Excel workbooks."""

    def __init__(self, file_path: str):
        """
        Initialize the chart creator with an Excel file.

        Args:
            file_path: Path to Excel file
        """
        self.file_path = file_path
        self.workbook = load_workbook(file_path)

    def add_bar_chart(
        self,
        sheet_name: str,
        data_range: str,
        categories_range: str,
        title: str,
        position: str = 'E2',
        width: float = 15,
        height: float = 10,
        chart_style: int = 2,
        stacked: bool = False
    ) -> None:
        """
        Add a bar chart to a worksheet.

        Args:
            sheet_name: Target worksheet
            data_range: Cell range for data (e.g., 'B2:D10')
            categories_range: Cell range for categories (e.g., 'A2:A10')
            title: Chart title
            position: Top-left cell for chart
            width, height: Chart dimensions in centimeters
            chart_style: Excel chart style (1-48)
            stacked: Whether bars should be stacked
        """
        sheet = self.workbook[sheet_name]

        # Create chart
        chart = BarChart()
        chart.title = title
        chart.style = chart_style
        chart.type = "col"

        if stacked:
            chart.grouping = "stacked"

        # Add data
        data = Reference(sheet, range_string=f"{sheet_name}!{data_range}")
        cats = Reference(sheet, range_string=f"{sheet_name}!{categories_range}")

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Add data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

        # Set size
        chart.width = width
        chart.height = height

        # Add to worksheet
        sheet.add_chart(chart, position)
        self.workbook.save(self.file_path)

    def add_line_chart(
        self,
        sheet_name: str,
        data_range: str,
        categories_range: str,
        title: str,
        position: str = 'E2',
        width: float = 15,
        height: float = 10,
        chart_style: int = 2,
        markers: bool = True
    ) -> None:
        """
        Add a line chart to a worksheet.

        Args:
            sheet_name: Target worksheet
            data_range: Cell range for data (e.g., 'B2:D10')
            categories_range: Cell range for categories (e.g., 'A2:A10')
            title: Chart title
            position: Top-left cell for chart
            width, height: Chart dimensions in centimeters
            chart_style: Excel chart style (1-48)
            markers: Whether to show markers
        """
        sheet = self.workbook[sheet_name]

        # Create chart
        chart = LineChart()
        chart.title = title
        chart.style = chart_style

        if markers:
            chart.marker = Marker(symbol='circle')

        # Add data
        data = Reference(sheet, range_string=f"{sheet_name}!{data_range}")
        cats = Reference(sheet, range_string=f"{sheet_name}!{categories_range}")

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Set size
        chart.width = width
        chart.height = height

        # Add to worksheet
        sheet.add_chart(chart, position)
        self.workbook.save(self.file_path)

    def add_pie_chart(
        self,
        sheet_name: str,
        data_range: str,
        categories_range: str,
        title: str,
        position: str = 'E2',
        width: float = 15,
        height: float = 10,
        chart_style: int = 2
    ) -> None:
        """
        Add a pie chart to a worksheet.

        Args:
            sheet_name: Target worksheet
            data_range: Cell range for data (e.g., 'B2:B10')
            categories_range: Cell range for categories (e.g., 'A2:A10')
            title: Chart title
            position: Top-left cell for chart
            width, height: Chart dimensions in centimeters
            chart_style: Excel chart style (1-48)
        """
        sheet = self.workbook[sheet_name]

        # Create chart
        chart = PieChart()
        chart.title = title
        chart.style = chart_style

        # Add data
        data = Reference(sheet, range_string=f"{sheet_name}!{data_range}")
        cats = Reference(sheet, range_string=f"{sheet_name}!{categories_range}")

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Add data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True

        # Set size
        chart.width = width
        chart.height = height

        # Add to worksheet
        sheet.add_chart(chart, position)
        self.workbook.save(self.file_path)

    def add_scatter_chart(
        self,
        sheet_name: str,
        x_data_range: str,
        y_data_range: str,
        title: str,
        position: str = 'E2',
        width: float = 15,
        height: float = 10,
        chart_style: int = 2
    ) -> None:
        """
        Add a scatter chart to a worksheet.

        Args:
            sheet_name: Target worksheet
            x_data_range: Cell range for X data (e.g., 'A2:A10')
            y_data_range: Cell range for Y data (e.g., 'B2:B10')
            title: Chart title
            position: Top-left cell for chart
            width, height: Chart dimensions in centimeters
            chart_style: Excel chart style (1-48)
        """
        sheet = self.workbook[sheet_name]

        # Create chart
        chart = ScatterChart()
        chart.title = title
        chart.style = chart_style

        # Add X data
        x_values = Reference(sheet, range_string=f"{sheet_name}!{x_data_range}")

        # Add Y data and create series
        y_values = Reference(sheet, range_string=f"{sheet_name}!{y_data_range}")
        series = Series(y_values, x_values, title="")

        chart.series.append(series)

        # Set size
        chart.width = width
        chart.height = height

        # Add to worksheet
        sheet.add_chart(chart, position)
        self.workbook.save(self.file_path)

    def save(self) -> None:
        """Save the workbook."""
        self.workbook.save(self.file_path)

    def close(self) -> None:
        """Close the workbook."""
        self.workbook.close()


class DataVisualizer:
    """
    Class for creating advanced visualizations using matplotlib and seaborn,
    and embedding them in Excel reports.
    """

    def __init__(self, save_dir: str = "output"):
        """
        Initialize visualizer with save directory.

        Args:
            save_dir: Directory to save generated images
        """
        self.save_dir = save_dir
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

    def create_time_series_plot(
        self,
        df: pd.DataFrame,
        date_col: str,
        value_cols: List[str],
        title: str = "Time Series Analysis",
        figsize: Tuple[int, int] = (12, 6),
        save_as: str = "time_series.png"
    ) -> str:
        """
        Create a time series plot.

        Args:
            df: Pandas DataFrame
            date_col: Column name containing dates
            value_cols: Column names for values to plot
            title: Plot title
            figsize: Figure size (width, height) in inches
            save_as: Filename to save plot

        Returns:
            Path to saved image
        """
        # Ensure date column is datetime
        df = df.copy()
        df[date_col] = pd.to_datetime(df[date_col])

        # Create plot
        plt.figure(figsize=figsize)
        for col in value_cols:
            plt.plot(df[date_col], df[col], marker='o', linestyle='-', label=col)

        plt.title(title)
        plt.xlabel('')
        plt.ylabel('Value')
        plt.grid(True, alpha=0.3)
        plt.legend()
        plt.xticks(rotation=45)
        plt.tight_layout()

        # Save plot
        save_path = os.path.join(self.save_dir, save_as)
        plt.savefig(save_path, dpi=300)
        plt.close()

        return save_path

    def create_correlation_heatmap(
        self,
        df: pd.DataFrame,
        numeric_cols: Optional[List[str]] = None,
        title: str = "Correlation Matrix",
        figsize: Tuple[int, int] = (10, 8),
        save_as: str = "correlation.png"
    ) -> str:
        """
        Create a correlation heatmap.

        Args:
            df: Pandas DataFrame
            numeric_cols: List of numeric columns to include (None for all numeric)
            title: Plot title
            figsize: Figure size (width, height) in inches
            save_as: Filename to save plot

        Returns:
            Path to saved image
        """
        # Select numeric columns if not specified
        df_corr = df.copy()
        if numeric_cols:
            df_corr = df_corr[numeric_cols]
        else:
            df_corr = df_corr.select_dtypes(include=['number'])

        # Calculate correlation
        corr_matrix = df_corr.corr()

        # Create heatmap
        plt.figure(figsize=figsize)
        mask = np.triu(np.ones_like(corr_matrix))
        sns.heatmap(
            corr_matrix,
            mask=mask,
            annot=True,
            fmt=".2f",
            cmap="coolwarm",
            vmin=-1,
            vmax=1,
            center=0,
            square=True,
            linewidths=0.5,
            cbar_kws={"shrink": 0.8}
        )

        plt.title(title)
        plt.tight_layout()

        # Save plot
        save_path = os.path.join(self.save_dir, save_as)
        plt.savefig(save_path, dpi=300)
        plt.close()

        return save_path

    def create_distribution_plots(
        self,
        df: pd.DataFrame,
        numeric_cols: List[str],
        title: str = "Distribution Analysis",
        figsize: Tuple[int, int] = (12, 8),
        save_as: str = "distributions.png"
    ) -> str:
        """
        Create distribution plots for multiple columns.

        Args:
            df: Pandas DataFrame
            numeric_cols: List of numeric columns to include
            title: Plot title
            figsize: Figure size (width, height) in inches
            save_as: Filename to save plot

        Returns:
            Path to saved image
        """
        # Handle case with only one column
        if len(numeric_cols) == 1:
            plt.figure(figsize=figsize)
            sns.histplot(df[numeric_cols[0]], kde=True)
            plt.title(f"{numeric_cols[0]} Distribution")
            plt.xlabel(numeric_cols[0])
            plt.ylabel("Count")
            plt.tight_layout()

            # Save plot
            save_path = os.path.join(self.save_dir, save_as)
            plt.savefig(save_path, dpi=300)
            plt.close()

            return save_path

        # Determine grid dimensions
        n_cols = len(numeric_cols)
        n_rows = (n_cols + 1) // 2  # Ceiling division

        # Create subplots
        fig, axes = plt.subplots(n_rows, 2, figsize=figsize)
        fig.suptitle(title, fontsize=16)

        # Flatten axes array for easy iteration
        if n_rows > 1:
            axes = axes.flatten()

        # Create distribution plots
        for i, col in enumerate(numeric_cols):
            if i < len(axes):
                ax = axes[i] if n_rows > 1 else axes[i % 2]
                sns.histplot(df[col], kde=True, ax=ax)
                ax.set_title(f"{col} Distribution")
                ax.set_xlabel(col)
                ax.set_ylabel("Count")

        # Hide unused subplots
        if n_rows > 1:
            for j in range(n_cols, len(axes)):
                axes[j].set_visible(False)

        plt.tight_layout()
        plt.subplots_adjust(top=0.9)

        # Save plot
        save_path = os.path.join(self.save_dir, save_as)
        plt.savefig(save_path, dpi=300)
        plt.close()

        return save_path

    def create_box_plots(
        self,
        df: pd.DataFrame,
        numeric_cols: List[str],
        category_col: Optional[str] = None,
        title: str = "Box Plot Analysis",
        figsize: Tuple[int, int] = (12, 8),
        save_as: str = "boxplots.png"
    ) -> str:
        """
        Create box plots for multiple columns.

        Args:
            df: Pandas DataFrame
            numeric_cols: List of numeric columns to include
            category_col: Optional column to group by
            title: Plot title
            figsize: Figure size (width, height) in inches
            save_as: Filename to save plot

        Returns:
            Path to saved image
        """
        # Melt the dataframe for easier plotting
        if category_col:
            id_vars = [category_col]
            melt_df = df.melt(id_vars=id_vars, value_vars=numeric_cols,
                               var_name='Variable', value_name='Value')

            # Create plot
            plt.figure(figsize=figsize)
            sns.boxplot(data=melt_df, x='Variable', y='Value', hue=category_col)
        else:
            melt_df = df.melt(value_vars=numeric_cols, var_name='Variable', value_name='Value')

            # Create plot
            plt.figure(figsize=figsize)
            sns.boxplot(data=melt_df, x='Variable', y='Value')

        plt.title(title)
        plt.xticks(rotation=45)
        plt.tight_layout()

        # Save plot
        save_path = os.path.join(self.save_dir, save_as)
        plt.savefig(save_path, dpi=300)
        plt.close()

        return save_path


def embed_images_in_excel(excel_file: str, image_paths: List[str], sheet_name: str = 'Visualizations') -> None:
    """
    Embed images into an Excel workbook.

    Args:
        excel_file: Path to Excel file
        image_paths: List of paths to images to embed
        sheet_name: Name of the sheet to add images to
    """
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image

    # Load workbook
    wb = load_workbook(excel_file)

    # Create sheet if it doesn't exist
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    # Add each image
    row = 1
    for img_path in image_paths:
        img = Image(img_path)

        # Resize image if too large
        if img.width > 800:
            ratio = img.height / img.width
            img.width = 800
            img.height = int(800 * ratio)

        # Add to worksheet
        ws.add_image(img, f'A{row}')

        # Increment row for next image (with padding)
        row += (img.height // 20) + 2

    # Save workbook
    wb.save(excel_file)

    return excel_file